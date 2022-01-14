import ROOT
import math
import numpy as np
from art.morisot_2p0 import Morisot_2p0
import sys,os
from openpyxl import Workbook
import csv

# Initialize painter
myPainter = Morisot_2p0()
myPainter.colourpalette.setColourPalette("notSynthwave")
myPainter.labelType = 4
myPainter.savePDF = True
myPainter.luminosity = -1
myPainter.CME = -1
# 0 Just ATLAS    
# 1 "Preliminary"
# 2 "Internal"
# 3 "Simulation Preliminary"
# 4 "Simulation Internal"
# 5 "Simulation"
# 6 "Work in Progress"

# Prediction for electron destinations (multiple scattering): goudsmit-saunderson

# Standard
infilename = "../results_{0}micron_1e8events.root"
#infilename = "../results_{0}micron_5e10events_neutron.root"

# Only one particle?
#useOnly = "neutron"
useOnly = ""

# Additional tag?
tag = ""

# Neutron rebinning factor
rebin_neutrons = 5

def compute_rms_1D(hist) :
    sum2 = 0
    for bin in range(hist.GetNbinsX()+1) :
        sum2 += hist.GetBinContent(bin) * hist.GetBinCenter(bin)**2
    RMS = math.sqrt(sum2/hist.Integral())
    return RMS

def compute_rms_2D(hist) :
    sum2 = 0
    for binx in range(hist.GetNbinsX()+1) :
        for biny in range(hist.GetNbinsY()+1) :
          radius2 = hist.GetXaxis().GetBinCenter(binx)**2 + hist.GetYaxis().GetBinCenter(biny)**2
          sum2 += hist.GetBinContent(binx,biny) * radius2
    RMS = math.sqrt(sum2/hist.Integral())
    return RMS

savehists = {}
particle_dict = {"eminus" : {"full" : "Electrons", "short" : "e-"},
                 "eplus" : {"full" : "Positrons", "short" : "e+"},
                 "gamma" : {"full" : "Photons", "short" : "#gamma"},
                 "neutron" : {"full" : "Neutrons", "short" : "n"} }

# Pick up theory curves for e- multiple scattering
intheoryfile = ROOT.TFile.Open("../theory_curves.root","READ")
theory_curves = {}
for thickness in 1, 5, 10 :
  thisdict = {}
  multiple_scattering_DCS = intheoryfile.Get("MDCS_{0}micron".format(thickness))
  thisdict["rad"] = multiple_scattering_DCS
  multiple_scattering_DCS_deg = intheoryfile.Get("MDCS_{0}micron_deg".format(thickness))
  thisdict["deg"] = multiple_scattering_DCS_deg
  theory_curves[thickness] = thisdict
integrals_multiple_scattering = intheoryfile.Get("integrals_MDCS")
Widths_fromTDRFormula = intheoryfile.Get("Widths_fromTDRFormula")
Widths_fromTF1 = intheoryfile.Get("Widths_fromTF1")
intheoryfile.Close()

# Now get the main plots.
compareThickness = True
for thickness in 1, 5, 10 :

  thisThickness = {}
  iThickness = [1, 5, 10].index(thickness)

  thisname = infilename.format(thickness)
  print("Checking file",thisname)
  if not os.path.isfile(thisname) :
      print("No such file, continuing")
      compareThickness = False
      continue

  infile = ROOT.TFile.Open(thisname)

  for particle in ["eminus", "eplus", "gamma", "neutron"] :

    if (useOnly and particle not in useOnly) : continue

    particle_full = particle_dict[particle]["full"]
    particle_short = particle_dict[particle]["short"]

    # Normalised scattering angle
    polar_scattering = infile.Get("angle_polar_{0}".format(particle))
    polar_scattering.SetDirectory(0)
    polar_scattering_norm = infile.Get("angle_polar_{0}_norm".format(particle))
    polar_scattering_norm.SetDirectory(0)
    polar_scattering_norm.SetName(polar_scattering_norm.GetName()+"_{0}micron".format(thickness))

    # Normalised scattering angle, axis in deg
    polar_scattering_deg = infile.Get("angle_polar_deg_{0}".format(particle))
    polar_scattering_deg.SetDirectory(0)
    polar_scattering_deg_norm = infile.Get("angle_polar_deg_{0}_norm".format(particle))
    polar_scattering_deg_norm.SetDirectory(0)
    polar_scattering_deg_norm.SetName(polar_scattering_deg_norm.GetName()+"_{0}micron".format(thickness))

    # Rebin neutrons: low stats
    if ("neutron" in particle) :
        polar_scattering.Rebin(rebin_neutrons)
        polar_scattering_norm.Rebin(rebin_neutrons)
        polar_scattering_deg.Rebin(rebin_neutrons)
        polar_scattering_deg_norm.Rebin(rebin_neutrons)

    # Save
    thisThickness["polar_scattering_{0}".format(particle)] = polar_scattering
    thisThickness["polar_scattering_norm_{0}".format(particle)] = polar_scattering_norm    
    thisThickness["polar_scattering_deg_{0}".format(particle)] = polar_scattering_deg
    thisThickness["polar_scattering_deg_norm_{0}".format(particle)] = polar_scattering_deg_norm

    # Make individual plots by thickness
    # Radians, unnormalised
    yhigh = None
    if "neutron" in particle :
        yhigh = 10

    myPainter.drawOverlaidHistos([polar_scattering],data=None,signal_list=None,histos_labels=["Geant4 {0}".format(particle_short)],data_label="",xlabel="Polar angle [rad]",ylabel="{0} total".format(particle_full),plotname="plots/scattering_{0}_{1}microns".format(particle,thickness)+tag,doRatio=False,ratioName="",doBkgErr=False,logx=False,logy=True,nLegendColumns=1,extraLines=[],xlow=0,xhigh=3.14,ylow=None,yhigh=yhigh,useTrueEdges=True)
    myPainter.drawOverlaidHistos([polar_scattering],data=None,signal_list=None,histos_labels=["Geant4 {0}".format(particle_short)],data_label="",xlabel="Polar angle [rad]",ylabel="{0} total".format(particle_full),plotname="plots/scattering_{0}_{1}microns_zoom".format(particle,thickness)+tag,doRatio=False,ratioName="",doBkgErr=False,logx=False,logy=True,nLegendColumns=1,extraLines=[],xlow=0,xhigh=1.0,ylow=None,yhigh=yhigh,useTrueEdges=True)    
    # Radians, normalised
    myPainter.drawOverlaidHistos([polar_scattering_norm],data=None,signal_list=None,histos_labels=["Geant4 {0}".format(particle_short)],data_label="",xlabel="Polar angle [rad]",ylabel="{0} per unit solid angle".format(particle_full),plotname="plots/scattering_norm_{0}_{1}microns".format(particle,thickness)+tag,doRatio=False,ratioName="",doBkgErr=False,logx=False,logy=True,nLegendColumns=1,extraLines=[],xlow=0,xhigh=3.14,ylow=None,yhigh=yhigh,useTrueEdges=True)
    myPainter.drawOverlaidHistos([polar_scattering_norm],data=None,signal_list=None,histos_labels=["Geant4 {0}".format(particle_short)],data_label="",xlabel="Polar angle [rad]",ylabel="{0} per unit solid angle".format(particle_full),plotname="plots/scattering_norm_{0}_{1}microns_zoom".format(particle,thickness)+tag,doRatio=False,ratioName="",doBkgErr=False,logx=False,logy=True,nLegendColumns=1,extraLines=[],xlow=0,xhigh=1.0,ylow=None,yhigh=yhigh,useTrueEdges=True)
    # Degrees, unnormalised
    myPainter.drawOverlaidHistos([polar_scattering_deg],data=None,signal_list=None,histos_labels=["Geant4 {0}".format(particle_short)],data_label="",xlabel="Polar angle [deg]",ylabel="{0} total".format(particle_full),plotname="plots/scattering_deg_{0}_{1}microns".format(particle,thickness)+tag,doRatio=False,ratioName="",doBkgErr=False,logx=False,logy=True,nLegendColumns=1,extraLines=[],xlow=0,xhigh=180,ylow=None,yhigh=yhigh,useTrueEdges=True)
    myPainter.drawOverlaidHistos([polar_scattering_deg],data=None,signal_list=None,histos_labels=["Geant4 {0}".format(particle_short)],data_label="",xlabel="Polar angle [deg]",ylabel="{0} total".format(particle_full),plotname="plots/scattering_deg_{0}_{1}microns_zoom".format(particle,thickness)+tag,doRatio=False,ratioName="",doBkgErr=False,logx=False,logy=True,nLegendColumns=1,extraLines=[],xlow=0,xhigh=40,ylow=None,yhigh=yhigh,useTrueEdges=True)   
    # Degrees, normalised
    myPainter.drawOverlaidHistos([polar_scattering_deg_norm],data=None,signal_list=None,histos_labels=["Geant4 {0}".format(particle_short)],data_label="",xlabel="Polar angle [deg]",ylabel="{0} per unit solid angle".format(particle_full),plotname="plots/scattering_deg_norm_{0}_{1}microns".format(particle,thickness)+tag,doRatio=False,ratioName="",doBkgErr=False,logx=False,logy=True,nLegendColumns=1,extraLines=[],xlow=0,xhigh=180,ylow=None,yhigh=yhigh,useTrueEdges=True)
    myPainter.drawOverlaidHistos([polar_scattering_deg_norm],data=None,signal_list=None,histos_labels=["Geant4 {0}".format(particle_short)],data_label="",xlabel="Polar angle [deg]",ylabel="{0} per unit solid angle".format(particle_full),plotname="plots/scattering_deg_norm_{0}_{1}microns_zoom".format(particle,thickness)+tag,doRatio=False,ratioName="",doBkgErr=False,logx=False,logy=True,nLegendColumns=1,extraLines=[],xlow=0,xhigh=40,ylow=None,yhigh=yhigh,useTrueEdges=True) 

    # Now normalize to match theory curves and plot.
    # Only this if e-
    if "eminus" in particle :
        # Should already be fitted to the distribution
        rad_DCS = theory_curves[thickness]["rad"]
        myPainter.drawHistsWithTF1s([polar_scattering_norm],[rad_DCS],as_data=False,match_colours=True,hist_labels=["Geant4 e-"],func_labels=["Moliere DCS"],xlabel="Polar scattering angle [rad.]",ylabel="Arbitrary units",plotname="plots/scattering_w_theory_{0}microns".format(thickness)+tag,doRatios=False,ratioName="",doErrors=False,logx=False,logy=True,nLegendColumns=1,extraLines=[],xlow=0,xhigh=1.0,ylow=None,yhigh=1e3)
        deg_DCS = theory_curves[thickness]["deg"]
        myPainter.drawHistsWithTF1s([polar_scattering_deg_norm],[deg_DCS],as_data=False,match_colours=True,hist_labels=["Geant4 e-"],func_labels=["Moliere DCS"],xlabel="Polar scattering angle [deg.]",ylabel="Arbitrary units",plotname="plots/scattering_w_theory_deg_{0}microns".format(thickness)+tag,doRatios=False,ratioName="",doErrors=False,logx=False,logy=True,nLegendColumns=1,extraLines=[],xlow=0,xhigh=90.,ylow=None,yhigh=1e3)            

    # Save positions and momenta
    for histname in ["momentum_x_{0}".format(particle), "momentum_y_{0}".format(particle), "momentum_z_{0}".format(particle), "position_x_{0}".format(particle), "position_y_{0}".format(particle), "position_z_{0}".format(particle)] :
        thishist = infile.Get(histname)
        thishist.SetDirectory(0)
        thisThickness[histname] = thishist

    # p
    momentum = infile.Get("momentum_{0}".format(particle))
    momentum.SetDirectory(0)
    if "neutron" in particle : momentum.Rebin(rebin_neutrons)
    thisThickness["momentum_{0}".format(particle)] = momentum
    myPainter.drawOverlaidHistos([momentum],data=None,signal_list=None,histos_labels=["Geant4 {0}".format(particle_short)],data_label="",xlabel="Momentum [MeV]",ylabel="Number of {0}".format(particle_full),plotname="plots/momentum_{0}_{1}microns".format(particle,thickness)+tag,doRatio=False,ratioName="",doBkgErr=False,logx=False,logy=True,nLegendColumns=1,extraLines=[],xlow=0,xhigh=None,ylow=None,yhigh=None,useTrueEdges=True)

    # Energy (not representative for neutrons)
    energy = infile.Get("energy_{0}".format(particle))
    energy.SetDirectory(0)
    thisThickness["energy_{0}".format(particle)] = energy
    myPainter.drawOverlaidHistos([energy],data=None,signal_list=None,histos_labels=["Geant4 {0}".format(particle_short)],data_label="",xlabel="Energy [MeV]",ylabel="Number of {0}".format(particle_full),plotname="plots/energy_{0}_{1}microns".format(particle,thickness)+tag,doRatio=False,ratioName="",doBkgErr=False,logx=False,logy=True,nLegendColumns=1,extraLines=[],xlow=0,xhigh=None,ylow=None,yhigh=None,useTrueEdges=True)

    # Collect and save 2d histograms
    momentum_2d = infile.Get("momentum_xy_{0}".format(particle))
    momentum_2d.SetDirectory(0)
    thisThickness["momentum_xy_{0}".format(particle)] = momentum_2d
    position_2d = infile.Get("position_xy_{0}".format(particle))
    position_2d.SetDirectory(0)
    thisThickness["position_xy_{0}".format(particle)] = position_2d

    # Plot for visual fun
    myPainter.draw2DHist(momentum_2d,"plots/momentum_xy_{0}_{1}microns".format(particle,thickness)+tag,"px [MeV]",-30,30,"py [MeV]",-30,30,particle_full,logz=True)
    myPainter.draw2DHist(position_2d,"plots/position_xy_{0}_{1}microns".format(particle,thickness)+tag,"x [mm]",-1000,1000,"y [mm]",-1000,1000,particle_full,logz=True)

  savehists[thickness] = thisThickness

  # Make stacks of output particles within thickness. Only if doing all particles
  if (useOnly) : continue
  to_stack = [thisThickness["polar_scattering_norm_neutron"],thisThickness["polar_scattering_norm_eplus"],thisThickness["polar_scattering_norm_gamma"],thisThickness["polar_scattering_norm_eminus"]]
  myPainter.drawStackedHistos(to_stack,data=None,signal_list=None,stack_labels=["Neutrons","Positrons","Photons","Electrons"],data_label="",xlabel="Polar angle [rad]",ylabel="Norm. particles per rad^{2}",plotname="plots/stacked_particles_norm_{0}microns".format(thickness)+tag,doRatio=False,ratioName="",doBkgErr=False,logx=False,logy=True,nLegendColumns=1,extraLines=[],xlow=0,xhigh=1,ylow=None,yhigh=None)

  to_stack = [thisThickness["polar_scattering_neutron"],thisThickness["polar_scattering_eplus"],thisThickness["polar_scattering_gamma"],thisThickness["polar_scattering_eminus"]]
  myPainter.drawStackedHistos(to_stack,data=None,signal_list=None,stack_labels=["Neutrons","Positrons","Photons","Electrons"],data_label="",xlabel="Polar angle [rad]",ylabel="Total particles",plotname="plots/stacked_particles_{0}microns".format(thickness)+tag,doRatio=False,ratioName="",doBkgErr=False,logx=False,logy=True,nLegendColumns=1,extraLines=[],xlow=0,xhigh=1,ylow=None,yhigh=None)

  to_stack = [thisThickness["polar_scattering_deg_norm_neutron"],thisThickness["polar_scattering_deg_norm_eplus"],thisThickness["polar_scattering_deg_norm_gamma"],thisThickness["polar_scattering_deg_norm_eminus"]]
  myPainter.drawStackedHistos(to_stack,data=None,signal_list=None,stack_labels=["Neutrons","Positrons","Photons","Electrons"],data_label="",xlabel="Polar angle [deg]",ylabel="Norm. particles per deg^{2}",plotname="plots/stacked_particles_deg_norm_{0}microns".format(thickness)+tag,doRatio=False,ratioName="",doBkgErr=False,logx=False,logy=True,nLegendColumns=1,extraLines=[],xlow=0,xhigh=90,ylow=None,yhigh=None)

  to_stack = [thisThickness["polar_scattering_deg_neutron"],thisThickness["polar_scattering_deg_eplus"],thisThickness["polar_scattering_deg_gamma"],thisThickness["polar_scattering_deg_eminus"]]
  myPainter.drawStackedHistos(to_stack,data=None,signal_list=None,stack_labels=["Neutrons","Positrons","Photons","Electrons"],data_label="",xlabel="Polar angle [deg]",ylabel="Total particles",plotname="plots/stacked_particles_deg_{0}microns".format(thickness)+tag,doRatio=False,ratioName="",doBkgErr=False,logx=False,logy=True,nLegendColumns=1,extraLines=[],xlow=0,xhigh=90,ylow=None,yhigh=None)

# Close input file
infile.Close()

# Make plots comparing thicknesses, if we have all of them
if (compareThickness) :
    histos_list = [savehists[1]["polar_scattering_deg_eminus"],savehists[5]["polar_scattering_deg_eminus"],savehists[10]["polar_scattering_deg_eminus"]]
    myPainter.drawOverlaidHistos(histos_list,data=None,signal_list=None,histos_labels=["1 #mum foil","5 #mum foil","10 #mum foil"],data_label="",xlabel="Polar angle [deg]",ylabel="Total electrons",plotname="plots/electron_scattering_deg_compareThickness".format(thickness)+tag,doRatio=False,ratioName="",doBkgErr=False,logx=False,logy=True,nLegendColumns=1,extraLines=[],xlow=0,xhigh=45,ylow=None,yhigh=None)
    histos_list = [savehists[1]["polar_scattering_deg_norm_eminus"],savehists[5]["polar_scattering_deg_norm_eminus"],savehists[10]["polar_scattering_deg_norm_eminus"]]
    myPainter.drawOverlaidHistos(histos_list,data=None,signal_list=None,histos_labels=["1 #mum foil","5 #mum foil","10 #mum foil"],data_label="",xlabel="Polar angle [deg]",ylabel="Norm. electrons per deg^{2}",plotname="plots/electron_scattering_deg_norm_compareThickness".format(thickness)+tag,doRatio=False,ratioName="",doBkgErr=False,logx=False,logy=True,nLegendColumns=1,extraLines=[],xlow=0,xhigh=45,ylow=None,yhigh=None)

# Save histograms that I created here, so I can look at them further if desired
outfile = ROOT.TFile.Open("plots.root","RECREATE")
outfile.cd()
for thickness in savehists.keys() :
  for name,hist in savehists[thickness].items() :
    hist.Write()
outfile.Close()

# Now save them into a spreadsheet version.
# We'll do a new sheet per thickness, for ease.
keylist = sorted(savehists.keys())
for thickness in savehists.keys() :
  # Workbook for saving in excel format
  wb = Workbook()
  for index,name in enumerate(savehists[thickness].keys()) :

    # Do this only if not 2D:
    if "_xy_" in name : continue

    # TEMP FIXME: do this only if not the polar angle
    if "polar_scattering" not in name or "eminus" not in name : 
      print("Skipping hist",name)
      continue
    elif "deg" in name :
      print("Skipping hist",name)
      continue
    else :
      print("Saving hist",name)

    hist = savehists[thickness][name]
    if len(name) > 31 :
      safename = name[:31]
    else :
      safename = name
    if index is 0 :
      ws = wb.active
      ws.title = safename
    else :
      ws = wb.create_sheet(safename)

    # Now fill.
    ws['A1'] = "Title: "+hist.GetName()
    ws['A2'] = "Bin number"
    ws['B2'] = "Lower edge"
    ws['C2'] = "Upper edge"
    ws['D2'] = "Center"
    ws['E2'] = "Bin content"
    rowindex = 3
    for bin in range(1, hist.GetNbinsX()+1) :
      ws['A{0}'.format(rowindex)] = bin
      ws['B{0}'.format(rowindex)] = hist.GetBinLowEdge(bin)
      ws['C{0}'.format(rowindex)] = hist.GetBinLowEdge(bin+1)
      ws['D{0}'.format(rowindex)] = hist.GetBinCenter(bin)
      ws['E{0}'.format(rowindex)] = hist.GetBinContent(bin)
      rowindex += 1

  wb.save("histograms_{0}micronfoil.xlsx".format(thickness))

  # 2D hists are too large for workbooks. Save as .npy.
  for index,name in enumerate(savehists[thickness].keys()) :

    # Do this only if not 2D - those will be separate:
    if "_xy_" not in name : continue
    # Skip momentum and anything that's not electron
    if "momentum" in name : continue
    if "eminus" not in name : continue
    hist = savehists[thickness][name]
    
    # Checking info on this
    print("Hist",name,"has nbins",hist.GetNbinsX()*hist.GetNbinsY())

    # First row labels the columns.
    label_row = ["Bin number",
                "Lower edge (x) [mm]",
                "Upper edge (x) [mm]",
                "Lower edge (y) [mm]",
                "Upper edge (y) [mm]",
                "Bin content"]
    # Here we want to include overflow bins
    with open('{0}_{1}microntarget.csv'.format(name,thickness), 'w', newline='') as csvfile:
      filewriter = csv.writer(csvfile, delimiter='\t',
                            quotechar='|', quoting=csv.QUOTE_MINIMAL)
      filewriter.writerow(label_row)
      for binx in range(0, hist.GetNbinsX()+2) :
        for biny in range(0,hist.GetNbinsY()+2) :
          newrow = []
          newrow.append(hist.GetBin(binx,biny))
          newrow.append(hist.GetXaxis().GetBinLowEdge(binx))
          newrow.append(hist.GetXaxis().GetBinLowEdge(binx+1))
          newrow.append(hist.GetYaxis().GetBinLowEdge(biny))
          newrow.append(hist.GetYaxis().GetBinLowEdge(biny+1))
          newrow.append(hist.GetBinContent(binx,biny))
          filewriter.writerow(newrow)

# Beam spread results, electrons only - compare methods
if (useOnly and "eminus" not in useOnly) : sys.exit()
for thickness in 1, 5, 10 :
  # In case we did only one or two files
  if thickness not in savehists.keys() : continue

  # validate RMS calculation with one that is for sure OK
  test1 = savehists[thickness]["position_x_eminus"].GetRMS()
  test2 = compute_rms_1D(savehists[thickness]["position_x_eminus"])
  # print("Test RMS:",test1, test2) - validated
  test3 = savehists[thickness]["momentum_y_eminus"].GetRMS()
  test4 = compute_rms_1D(savehists[thickness]["momentum_y_eminus"])
  # print("Test RMS:",test3, test4) - validated

  # Widths evaluated from the TDR formula (PDG formula) and from the TF1s
  index = [1, 5, 10].index(thickness)
  estimate_formula = Widths_fromTDRFormula[index]
  estimate_TF1 = Widths_fromTF1[index]
  print("Width estimate from formula in TDR:",estimate_formula)
  print("Width estimate from TF1:",estimate_TF1)

  # Not using ROOT's RMS since it compares to the mean along the x axis and I want to compare to zero
  # RMS of normalised distribution
  RMS_norm_dist = compute_rms_1D(savehists[thickness]["polar_scattering_norm_eminus"])
  print("RMS from solid-angle-normalised Geant4 distribution {0:1.2g} = {1:1.2g} degrees".format(RMS_norm_dist,math.degrees(RMS_norm_dist)))
  # RMS by hand using bin center - non-normalised distribution. Pretty sure I do not want this
  #RMS_notnormalised = compute_rms_1D(savehists[thickness]["polar_scattering_eminus"])
  #print("RMS from unnormalised distribution {0:1.2g} = {1:1.2g} degrees".format(RMS_notnormalised,math.degrees(RMS_notnormalised)))

  # RMS of 2D distribution - position - and calculate from there
  RMS_position_2D = compute_rms_2D(savehists[thickness]["position_xy_eminus"])
  # Average using this is arctan RMS over 2 m
  RMS_angle_2D = math.atan(RMS_position_2D/2000.)
  print("RMS angle from 2D position distribution = {0:1.2g} = {1:1.2g} degrees".format(RMS_angle_2D,math.degrees(RMS_angle_2D)))

  # 1D (along each axis) of 2D distribution; how does that compare?
  RMS_xonly = math.atan(savehists[thickness]["position_xy_eminus"].GetRMS(1)/2000.)
  RMS_yonly = math.atan(savehists[thickness]["position_xy_eminus"].GetRMS(2)/2000.)
  print("For comparison, RMS along x only and y only give: {0:1.2g} {1:1.2g} = {2:1.2g} {3:1.2g} degrees".format(RMS_xonly,RMS_yonly,math.degrees(RMS_xonly),math.degrees(RMS_yonly)))
  print("from RMS x =",savehists[thickness]["position_xy_eminus"].GetRMS(1),"mm, RMS y =",savehists[thickness]["position_xy_eminus"].GetRMS(2),"mm")
